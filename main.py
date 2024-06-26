import json
from tkinter import *
from customtkinter import *
from datetime import date
import openpyxl
from openpyxl import Workbook
import random, ctypes, win32com.client
import xml.etree.ElementTree as ET
import requests



global_font = ("Helvetica", 25)

# Janela inicial
class UserWindow():
    def __init__(self):
        self.window = CTk()
        self.window.title("Consulta de componentes")
        self.window.geometry("500x500")
        self.window_input()
        self.window.mainloop()

    def window_input(self):
        CTkLabel(self.window, text="Consultar componentes:", font=global_font).pack()

        self.entry_pep = CTkEntry(self.window, 
                                  placeholder_text="Insira o PEP...", 
                                  font=global_font, 
                                  width=300)
        self.entry_pep.pack(pady=10)
        self.entry_mrp = CTkEntry(self.window, 
                                  placeholder_text="Insira o MRP...", 
                                  font=global_font, 
                                  width=300)
        self.entry_mrp.pack(pady=10)

        CTkButton(self.window, text="Consultar", 
                  command=self.get_input_values_after_click, 
                  font=global_font).pack(pady=30)
        ShowHistoryWindow(self.window)


    def get_input_values_after_click(self):
            self.pep = self.entry_pep.get()
            self.mrp = self.entry_mrp.get()

            if self.pep != "" or self.mrp != "":
                StartSap(self.pep, self.mrp)
    
# Janela dos resultados dos componentes extraídos do SAP
class StartSap():
    def __init__(self, pep, mrp):
        self.pep = pep
        self.mrp = mrp
        self.init_sap_proccess(self.pep, self.mrp)

        components = sap.data_collected_after_sap_proccess
        item_json = {f"{pep} - {mrp}": components}

        db.save_components(item_json)
        ShowItemWindow(f"{self.pep} - {self.mrp}")
               
    def init_sap_proccess(self, pep, mrp):
        sap.enter_coois(mrp, pep)
        sap.get_orders()
        sap.get_components()
        sap.get_almoxarifado()

# Janela do histórico de todos os peps já consultados
class ShowHistoryWindow():
    def __init__(self, window):
        self.window = CTkScrollableFrame(master=window, 
                                         border_width=1, 
                                         orientation="vertical", 
                                         fg_color="#3b8ed0", 
                                         width=300)
        self.window.pack(pady=1)
        CTkLabel(self.window, text="Histórico:", font=global_font).pack(pady=3)

        self.pep_hisotry_data = db.get_database()["consultas"]
        self.add_all_peps_to_window()
    
    def add_all_peps_to_window(self):
        for item in enumerate(self.pep_hisotry_data):
            pep_obj = item[1].keys()
            pep_number = list(pep_obj)[0]
            button = CTkButton(master=self.window, 
                               text=pep_number, 
                               command=lambda 
                               pep=pep_number: self.show_window(pep), 
                               font=global_font, 
                               corner_radius=0)
            button.pack()

    def show_window(self, pep_number):
        ShowItemWindow(f"{pep_number}")

# Janela dos componentes do pep selecionado no histórico
class ShowItemWindow():
    def __init__(self, pep):
        self.pep = pep
        self.description_width = 300

        self.row_height = 5 # tamanho das linhas da tabela

        self.show_window()
        self.pep_data = db.get_item_by_key(pep)
        self.write_all_peps_to_window()

        
    def show_window(self):
        self.window_width = 800
        self.window_height = 0

        self.window = CTk()
        self.window.title(f"PEP - {self.pep}")
        # self.window.geometry(f"{self.window_width}x{self.window_height}")

    def write_all_peps_to_window(self):
        self.row_qty = 0
        for line, item in enumerate(self.pep_data, start=1):
            self.add_row(item, line)

        btn_export = CTkButton(self.window, text="Exportar planilha", 
                               command=self.export_to_worksheet, 
                               font=global_font,
                               height=self.row_height)
        btn_export.grid(row=self.row_qty+1, column=0)
        # adicionar height na janela para o botao
        self.window_height += 50
        self.window.geometry(f"{self.window_width}x{self.window_height}")
        self.window.mainloop()

    def add_row(self, item, grid_row):
        text_field = CTkEntry(master=self.window, 
                              width=self.description_width, 
                              font=("Arial", 14), 
                              corner_radius=0)
        text_field.insert(0, "COMPONENTE")  
        text_field.grid(row=0, column=0)
        text_field = CTkEntry(
            master=self.window,
              width=100, 
              font=("Arial", 14), 
              corner_radius=0)
        text_field.insert(0, "MATERIAL")  
        text_field.grid(row=0, column=1)
        text_field = CTkEntry(master=self.window, 
                              width=100, 
                              font=("Arial", 14), 
                              corner_radius=0)
        text_field.insert(0, "QTD")  
        text_field.grid(row=0, column=2)
        text_field = CTkEntry(master=self.window, 
                              width=100, 
                              font=("Arial", 14), 
                              corner_radius=0)
        text_field.insert(0, "RS01")  
        text_field.grid(row=0, column=3)
        text_field = CTkEntry(master=self.window, 
                              width=100, 
                              font=("Arial", 14), 
                              corner_radius=0)
        text_field.insert(0, "RS02")  
        text_field.grid(row=0, column=4)
        text_field = CTkEntry(master=self.window, 
                              width=100, 
                              font=("Arial", 14), 
                              corner_radius=0)
        text_field.insert(0, "IA04")  
        text_field.grid(row=0, column=5)

        self.row_qty = 0
        for column, value in enumerate(item, start=0):
            desc_component = 0
            width=100
            if column == desc_component:
                width=self.description_width
            value = item[column]
            if value is None or value.strip() == "":
                value = 0
            else:
                value = value.split(",", 1)[0]
            text_field = CTkEntry(master=self.window, 
                                  width=width, 
                                  font=("Arial", 12), 
                                  corner_radius=0,
                                  height=self.row_height)
            text_field.insert(0, value)  
            text_field.grid(row=grid_row, column=column)
            self.row_qty = grid_row
            # deixa a altura da janela de acordo com a row_height
            self.window_height += self.row_height

    def export_to_worksheet(self):
        """ Cria uma planilha com a tabela dos dados do PEP
        - Salva no siretório de downloads
        - Estiliza a planilha com autofit
        """
        wb = Workbook()
        ws = wb.active
        today_date = date.today().strftime("%Y%m%d")
        ws.title = f"{today_date}"
        headers = ['Componente', 
                   'Material', 
                   'Necessário', 
                   'RS01', 
                   'RS02', 
                   'IA04']
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)
   
        for row_idx, row_data in enumerate(self.pep_data, start=2):
            for col_idx, cell_data in enumerate(row_data, start=1):
                if cell_data is not None:
                    try:
                        number_value = float(cell_data.replace(',', '.'))
                        ws.cell(row=row_idx, 
                                column=col_idx, 
                                value=number_value)
                    except ValueError: # é string
                        if cell_data.strip() == "":
                            cell_data = 0 
                        ws.cell(row=row_idx, 
                                column=col_idx, 
                                value=cell_data)
                else: # nulo, valor 0
                    ws.cell(row=row_idx, 
                            column=col_idx, 
                            value=0)
        # autofit do width
        for col_idx in range(1, 6): 
            max_length = 0
            iter_rows = ws.iter_rows(
                min_row=1, 
                max_row=len(self.pep_data) + 1, 
                min_col=col_idx, 
                max_col=col_idx)
            for row in iter_rows:
                for cell in row:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except TypeError:
                        pass
            ws.column_dimensions[
                ws.cell(row=1, column=col_idx).column_letter
            ].width = max_length + 2


        rnd_num = random.randint(0,999)
        file_name = f"consulta-{self.pep}-{today_date}{rnd_num}.xlsx"

        user_downloads_dir = os.path.expanduser("~/Downloads")
        dir_to_save = os.path.join(user_downloads_dir, file_name)
        wb.save(dir_to_save)
        wb.close()
        export_msg = f"Planilha {file_name} exportada em {user_downloads_dir}, espere a planilha abrir"
        ctypes.windll.user32.MessageBoxW(0, 
                                         export_msg, 
                                         "Exportar", 
                                         0
                                         )
        os.startfile(dir_to_save)

def set_text_before(column, value):
    match(column):
        case 3: return f"Total: {value}"
        case 4: return f"RS01: {value}"
        case 5: return f"RS02: {value}"
        case 6: return f"IA04: {value}"
        case _: return value

class Database:
    def __init__(self):
        self.db = self.get_database()

    def get_database(self):
        with open('database.json', 'r') as db:
            return json.load(db)
    
    def get_item_by_key(self, key):
        data = self.get_database()["consultas"]
        for consulta in data:
            if key in consulta:
                return consulta[key]
    
    def save_components(self, item):
         with open("database.json",'r+') as db:
            data = json.load(db)
            data["consultas"].append(item)
            db.seek(0)
            json.dump(data, db, indent = 4)

class SAP:
    def __init__(self):
        self.session = None
        self.connect_sap()
        self.orders = []
        self.materials = []
        self.components = []
        self.data_collected_after_sap_proccess = []

    def connect_sap(self):
        try:
            sapguiauto = win32com.client.GetObject("SAPGUI")
            application = sapguiauto.GetScriptingEngine
            connection = application.Children(0)
            self.session = connection.Children(0)
        except:
            print("Abra o SAP para rodar o programa!")
            exit

    def enter_coois(self, mrp, pep):
        self.session.findById("wnd[0]").maximize
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/ncoois"
        self.session.findById("wnd[0]").sendVKey(0)
        self.session.findById("wnd[0]/usr/ssub%_SUBSCREEN_TOPBLOCK:PPIO_ENTRY:1100/ctxtPPIO_ENTRY_SC1100-ALV_VARIANT").text = "/danielsson"
        self.session.findById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_DISPO-LOW").text = mrp
        self.session.findById("wnd[0]/usr/tabsTABSTRIP_SELBLOCK/tabpSEL_00/ssub%_SUBSCREEN_SELBLOCK:PPIO_ENTRY:1200/ctxtS_PROJN-LOW").text = pep
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()

    def get_orders(self):
        try:
            for line in range(50):
                order = self.session.findById("wnd[0]/usr/cntlCUSTOM/shellcont/shell/shellcont/shell").getCellValue(line, "AUFNR")
                self.orders.append(order)
        except:
            pass

    def get_components(self):
        for order in self.orders:
            url = f"http://10.1.75.70:82/wdc-integrations/APIop/?ordem={order}"
            response = requests.get(url)
            root = ET.fromstring(response.content)
            sequences = root.findall('.//ProductionOrder/Sequence')
            for sequence in sequences:
                materials = sequence.findall('.//Operation/MaterialInput')
                for material in materials:
                    material_elem = material.find(".//MaterialID")
                    material_text = material_elem.text
                    material_number = material_text[8:]
                    qty_elem = material.find(".//RequirementQuantity")
                    qty_with_dots = qty_elem.text
                    qty = qty_with_dots.split(".")[0]
                    desc_elem = material.find(".//MaterialDescription")
                    desc = desc_elem.text
                    op_elem = material.find(".//OperationID")
                    op = op_elem.text
                    if op == "1050":
                        self.materials.append(
                            [desc, material_number, qty]
                        )
            self.sum_values()

    def get_almoxarifado(self):
        for num, material in enumerate(self.materials, start=0):
            print(self.materials[num])
            try:
                qtd_rs01 = self.find_mmbe_material(material[1], "RS01")
                qtd_rs02 = self.find_mmbe_material(material[1], "RS02")
                qtd_ia04 = self.find_mmbe_material(material[1], "IA04")
                self.materials[num].append(qtd_rs01)
                self.materials[num].append(qtd_rs02)
                self.materials[num].append(qtd_ia04)
                
            except:
                pass
        self.data_collected_after_sap_proccess = self.materials

    def find_mmbe_material(self, material, almoxarifado):
        self.session.findById("wnd[0]/tbar[0]/okcd").text = "/nmmbe"
        self.session.findById("wnd[0]").sendVKey(0)
        self.session.findById("wnd[0]/usr/ctxtMS_MATNR-LOW").text = material
        self.session.findById("wnd[0]/usr/ctxtMS_LGORT-LOW").text = almoxarifado
        
        self.session.findById("wnd[0]/tbar[1]/btn[8]").press()
        try:
            almoxarife = self.session.findById("wnd[0]/usr/cntlCC_CONTAINER/shellcont/shell/shellcont[1]/shell[1]").getItemText("          4","&Hierarchy")
            almoxarife_2 = self.session.findById("wnd[0]/usr/cntlCC_CONTAINER/shellcont/shell/shellcont[1]/shell[1]").getItemText("          5","&Hierarchy")
    
        except:
            pass
        try:
            # As vezes é na linha 4
            if "RS01" in almoxarife or "RS02" in almoxarife or "IA04" in almoxarife:
                qtd = self.session.findById("wnd[0]/usr/cntlCC_CONTAINER/shellcont/shell/shellcont[1]/shell[1]").getItemText("          4","C          1")
                print(qtd)
                if qtd is not None or qtd != "":
                    return qtd
                else:
                    return "0"
                
            # As vezes é na linha 5
            if "RS01" in almoxarife_2 or "RS02" in almoxarife_2 or "IA04" in almoxarife_2:
                
                qtd = self.session.findById("wnd[0]/usr/cntlCC_CONTAINER/shellcont/shell/shellcont[1]/shell[1]").getItemText("          5","C          1")
                print(qtd)
                if qtd is not None or qtd != "":
                    return qtd
                else:
                    return "0"
    
        except:
            return "0"
        
    def reset_components_for_next_data(self):
        self.orders = []
        self.data_collected_after_sap_proccess = []


    def sum_values(self):
        summed_components = {}
        for component in self.materials:
            index = component[1]

            quantity = int(component[2].replace(',', '').strip())

            if index in summed_components:
                summed_components[index][2] = str(int(
                    summed_components[index][2].replace(',', '').strip()) + quantity
                    )
            else:
                summed_components[index] = component
        self.materials = list(summed_components.values())

try:
    sap = SAP()
    db = Database()
    UserWindow()
except:
    ctypes.windll.user32.MessageBoxW(0, 
                                    "Abra o SAP para rodar o programa!!!!!!!!", 
                                    "Aviso", 
                                    0
                                    )
# Exemplo para testes
# 120-2100844-22
# 100